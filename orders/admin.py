import io
from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
from .models import Cart, CartItem, PromoCode, Order, OrderItem, PaymentReceipt, ShippingZone


def export_orders_excel(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Orders'

    header_fill = PatternFill(start_color='FFB3C6', end_color='FFB3C6', fill_type='solid')
    header_font = Font(bold=True, color='4A1A2C')

    headers = [
        'Order #', 'Date', 'Customer Name', 'Phone', 'Email',
        'Governorate', 'City', 'Address',
        'Items', 'Subtotal (EGP)', 'Shipping (EGP)', 'Discount (EGP)', 'Total (EGP)',
        'Payment Method', 'Payment Status', 'Order Status',
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    orders = queryset.prefetch_related('items').order_by('-created_at')
    for order in orders:
        items_str = '; '.join(
            f'{i.quantity}x {i.product_name} ({i.price} EGP)' for i in order.items.all()
        )
        ws.append([
            order.order_number,
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            order.full_name,
            order.phone,
            order.email,
            order.governorate,
            order.city,
            order.address,
            items_str,
            float(order.subtotal),
            float(order.shipping_fee),
            float(order.discount_amount),
            float(order.total),
            order.get_payment_method_display(),
            order.get_payment_status_display(),
            order.get_status_display(),
        ])

    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ''))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="sheglow_orders.xlsx"'
    return response

export_orders_excel.short_description = 'Export selected orders to Excel'


@admin.register(ShippingZone)
class ShippingZoneAdmin(admin.ModelAdmin):
    list_display = ('governorate', 'shipping_fee', 'delivery_days', 'is_active')
    list_editable = ('shipping_fee', 'delivery_days', 'is_active')
    search_fields = ('governorate',)
    ordering = ('governorate',)


class CartItemInline(admin.TabularInline):
    model = CartItem
    extra = 0
    readonly_fields = ('product', 'quantity', 'subtotal')

    def subtotal(self, obj):
        return f'{obj.subtotal:,.0f} EGP'


@admin.register(Cart)
class CartAdmin(admin.ModelAdmin):
    list_display = ('__str__', 'item_count', 'subtotal_display', 'created_at', 'updated_at')
    list_filter = ('created_at',)
    search_fields = ('user__email', 'session_key')
    inlines = [CartItemInline]
    readonly_fields = ('created_at', 'updated_at')

    def item_count(self, obj):
        return obj.get_item_count()
    item_count.short_description = 'Items'

    def subtotal_display(self, obj):
        return f'{obj.get_subtotal():,.0f} EGP'
    subtotal_display.short_description = 'Subtotal'


@admin.register(PromoCode)
class PromoCodeAdmin(admin.ModelAdmin):
    list_display = ('code', 'discount_type', 'discount_value', 'min_order_amount',
                    'max_uses', 'times_used', 'is_active', 'valid_from', 'valid_until')
    list_filter = ('discount_type', 'is_active')
    search_fields = ('code',)
    list_editable = ('is_active',)
    readonly_fields = ('times_used',)


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 1
    fields = ('product', 'product_name', 'price', 'quantity', 'subtotal')
    readonly_fields = ('subtotal',)

    def subtotal(self, obj):
        if obj.pk:
            return f'{obj.subtotal:,.0f} EGP'
        return '—'
    subtotal.short_description = 'Subtotal'


class PaymentReceiptInline(admin.StackedInline):
    model = PaymentReceipt
    extra = 0
    readonly_fields = ('receipt_preview', 'uploaded_at')

    def receipt_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" style="max-height:150px;max-width:300px;">', obj.image.url)
        return '—'
    receipt_preview.short_description = 'Preview'


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ('order_number', 'full_name', 'governorate', 'total_display',
                    'payment_method', 'payment_status', 'status', 'created_at')
    list_filter = ('status', 'payment_method', 'payment_status', 'governorate', 'created_at')
    search_fields = ('order_number', 'full_name', 'email', 'phone')
    list_editable = ('status', 'payment_status')
    readonly_fields = ('order_number', 'created_at', 'updated_at')
    actions = [export_orders_excel]
    inlines = [OrderItemInline, PaymentReceiptInline]
    fieldsets = (
        ('Order Info', {
            'fields': ('order_number', 'user', 'status', 'created_at', 'updated_at')
        }),
        ('Shipping', {
            'fields': ('full_name', 'email', 'phone', 'address', 'city', 'governorate')
        }),
        ('Payment', {
            'fields': ('payment_method', 'payment_status')
        }),
        ('Financials', {
            'fields': ('subtotal', 'shipping_fee', 'promo_code', 'discount_amount', 'total')
        }),
    )

    def total_display(self, obj):
        return f'{obj.total:,.0f} EGP'
    total_display.short_description = 'Total'
    total_display.admin_order_field = 'total'

    def save_model(self, request, obj, form, change):
        if not change:
            # New manual order: auto-assign shipping fee from zone if not set
            from orders.views import get_shipping_fee
            if not obj.shipping_fee:
                obj.shipping_fee = get_shipping_fee(obj.governorate)
        super().save_model(request, obj, form, change)

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        obj = form.instance
        # Recalculate financials from actual order items
        items = obj.items.all()
        for item in items:
            # Auto-fill product_name and price from product if missing
            if item.product and not item.product_name:
                item.product_name = item.product.name
            if item.product and not item.price:
                item.price = item.product.effective_price
            if item.price and item.quantity:
                item.subtotal = item.price * item.quantity
            item.save(update_fields=['product_name', 'price', 'subtotal'])
            # Decrement stock only on creation
            if not change and item.product:
                from products.models import Product as Prod
                Prod.objects.filter(pk=item.product.pk).update(
                    stock=item.product.stock - item.quantity
                )
        # Recompute order totals
        subtotal = sum(i.subtotal for i in obj.items.all())
        obj.subtotal = subtotal
        if not obj.shipping_fee:
            from orders.views import get_shipping_fee
            obj.shipping_fee = get_shipping_fee(obj.governorate)
        obj.total = obj.subtotal + obj.shipping_fee - (obj.discount_amount or 0)
        obj.save(update_fields=['subtotal', 'total'])


@admin.register(PaymentReceipt)
class PaymentReceiptAdmin(admin.ModelAdmin):
    list_display = ('order', 'uploaded_at', 'verified', 'verified_by', 'receipt_preview')
    list_filter = ('verified',)
    search_fields = ('order__order_number',)
    list_editable = ('verified',)
    readonly_fields = ('uploaded_at', 'receipt_preview')

    def receipt_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" style="max-height:150px;max-width:300px;">', obj.image.url)
        return '—'
    receipt_preview.short_description = 'Receipt'
